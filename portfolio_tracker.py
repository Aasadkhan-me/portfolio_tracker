import datetime as dt
from pathlib import Path
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class PortfolioTracker:
    def __init__(self, holdings: dict):
        self.holdings = holdings
        self.prices = {}
        
    def fetch_live_prices(self):
        """Fetch current price and company name securely with data safeguards."""
        print(f"Fetching live prices for {len(self.holdings)} tickers...")
        for symbol in self.holdings:
            try:
                ticker = yf.Ticker(symbol)
                     # Safeguard against yfinance API delays or layout changes
                price = ticker.fast_info.get("lastPrice") or ticker.info.get("regularMarketPrice")
                name = ticker.info.get("shortName", symbol)
                
                if price:
                    self.prices[symbol] = {"price": round(price, 2), "name": name}
            except Exception as e:
                print(f"⚠️ Error pulling market data for {symbol}: {e}")

    def generate_excel_report(self, output_dir: str = "output"):
        """Compiles market data into a highly formatted, reactive Excel layout."""
        if not self.prices:
            print(" Execution halted: No price data available.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio Summary"

               # Apply stylized corporate headers

        headers = ["Ticker", "Company", "Quantity", "Price", "Market Value", "Allocation %"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")

               # Track explicitly where valid data rows print

        current_row = 2
        for symbol, qty in self.holdings.items():
            if symbol not in self.prices:
                continue
                
            data = self.prices[symbol]
            ws.append([symbol, data["name"], qty, data["price"]])
            
               # Formulate computations dynamically to prevent row shifting bugs

            ws.cell(row=current_row, column=5, value=f"=C{current_row}*D{current_row}")
            
               # Format display configurations

            ws.cell(row=current_row, column=4).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=5).number_format = "$#,##0.00"
            current_row += 1
 
               # Establish fixed relative spacing for totals

        total_row = current_row + 1
        
               # Backfill allocation formulas now that total_row target location is locked

        for row in range(2, current_row):
            ws.cell(row=row, column=6, value=f"=E{row}/$E${total_row}").number_format = "0.0%"

               # Set calculations for totals row

        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{current_row-1})")
        ws.cell(row=total_row, column=5).number_format = "$#,##0.00"
        ws.cell(row=total_row, column=6, value=1).number_format = "0.0%"

                # Set responsive style buffers
        for col, width in zip("ABCDEF", [12, 26, 12, 14, 16, 14]):
            ws.column_dimensions[col].width = width

                 # Execute platform-agnostic write routines
        save_path = Path(output_dir) / f"portfolio_{dt.date.today()}.xlsx"
        save_path.parent.mkdir(exist_ok=True)
        wb.save(save_path)
        print(f"📊 Live sheet generated at: {save_path}")


if __name__ == "__main__":
    
    # Define client data
    PORTFOLIO_DATA = {
        "AAPL": 25, "MSFT": 15, "GOOGL": 10, "AMZN": 12, "NVDA": 8,
    }
    
    # Run pipeline agent
    tracker = PortfolioTracker(PORTFOLIO_DATA)
    tracker.fetch_live_prices()
    tracker.generate_excel_report()
